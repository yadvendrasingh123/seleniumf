
import configparser
config = configparser.ConfigParser()
config.read("Utilities/.properties")
import openpyxl
from selenium import webdriver
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "excel1 demo1"
driver=webdriver.Chrome(executable_path="Driver/chromedriver.exe")
driver.get(config.get("Url2","base2_url"))
driver.implicitly_wait(5)

number_of_rows=driver.find_elements_by_xpath("//div[@class='tableFixHead']//tbody/tr")
i=1
for item in number_of_rows:
    sheet.cell(i,1).value =driver.find_element_by_xpath("//div[@class='tableFixHead']//tbody/tr["+str(i)+"]/td[1]").text
    sheet.cell(i,2).value = driver.find_element_by_xpath("//div[@class='tableFixHead']//tbody/tr["+str(i)+"]/td[2]").text
    sheet.cell(i,3).value = driver.find_element_by_xpath("//div[@class='tableFixHead']//tbody/tr["+str(i)+"]/td[3]").text
    sheet.cell(i,4).value = driver.find_element_by_xpath("//div[@class='tableFixHead']//tbody/tr["+str(i)+"]/td[4]").text
    i=i+1
workbook.save('hello.xlsx')